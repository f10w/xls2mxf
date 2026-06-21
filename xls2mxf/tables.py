"""Работа с траффик-листами: парсинг блоков, ID, даты, имена."""
import datetime as _dt
from pathlib import Path

import openpyxl

from .constants import HEADER_TEXT, EXT, DATE_RE
from .errors import AssemblyError


def _find_id_column(ws) -> int | None:
    """Ищет столбец с заголовком HEADER_TEXT. Возвращает индекс столбца (1-based)
    и строку шапки, либо None. Заголовок ищется в первых строках листа."""
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == HEADER_TEXT:
                return cell.column, cell.row
    return None



def extract_ids(xlsx_path: Path) -> set:
    """Извлекает целочисленные ID из столбца, найденного по заголовку 'ID ролика'.
    Не зависит от того, в каком столбце (F, J, ...) он находится."""
    ids = set()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        found = _find_id_column(ws)
        if not found:
            continue  # на этом листе нет колонки ID ролика
        col, header_row = found
        for row in ws.iter_rows(min_row=header_row + 1, min_col=col, max_col=col,
                                values_only=True):
            val = row[0]
            if val is None:
                continue
            if isinstance(val, str):
                s = val.strip()
                if not s.isdigit():
                    continue
                ids.add(int(s))
            elif isinstance(val, (int, float)) and float(val).is_integer():
                ids.add(int(val))
    wb.close()
    return ids



def find_xlsx_for_date(xlsx_dir: Path, ddmmyy: str) -> list:
    out = []
    for x in sorted(xlsx_dir.glob("*.xlsx")):
        if x.name.startswith("~$"):
            continue
        m = DATE_RE.search(x.stem)
        if m and m.group(1) == ddmmyy:
            out.append(x)
    return out



def read_id_column_raw(xlsx_path: Path, customlines: list) -> list:
    """Читает столбец 'ID ролика' СВЕРХУ ВНИЗ как есть: без сортировки и без
    удаления дубликатов. Каждый пропуск (пустые строки между блоками данных)
    заменяется ровно на три строки customlines. Пустоты до первого блока и
    после последнего отбрасываются.

    Возвращает список строк — готовый к построчной вставке (как столбец в Excel)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    found = _find_id_column(ws)
    if not found:
        wb.close()
        return []
    col, header_row = found

    # собираем "сырые" значения столбца ниже шапки: либо текст значения, либо None
    raw = []
    for row in ws.iter_rows(min_row=header_row + 1, min_col=col, max_col=col,
                            values_only=True):
        v = row[0]
        if v is None or (isinstance(v, str) and v.strip() == ""):
            raw.append(None)
        elif isinstance(v, float) and v.is_integer():
            raw.append(str(int(v)))
        else:
            raw.append(str(v).strip())
    wb.close()

    # обрезаем ведущие и хвостовые пустоты
    start, end = 0, len(raw)
    while start < end and raw[start] is None:
        start += 1
    while end > start and raw[end - 1] is None:
        end -= 1
    raw = raw[start:end]

    # любой внутренний пропуск (одна или несколько подряд пустых) -> 3 кастома
    result = []
    i = 0
    n = len(raw)
    while i < n:
        if raw[i] is not None:
            result.append(raw[i])
            i += 1
        else:
            # проматываем всю пачку пустых
            while i < n and raw[i] is None:
                i += 1
            result.extend(customlines)
    return result



def parse_blocks(xlsx_path: Path) -> list:
    """Режет траффик-лист на блоки. Возвращает список словарей:
       {time: 'HH:MM', ids: [int,...], itogo: int|None}.
    Блок = подряд идущие строки с числом в столбце ID; завершается строкой,
    где ID пуст. Время берётся из столбца A первой строки блока. ИТОГО — из
    столбца 'Хрон.' (E) в строке, где D == 'ИТОГО'."""
    import datetime as _dt
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    found = _find_id_column(ws)
    if not found:
        wb.close()
        raise AssemblyError(f"В таблице {xlsx_path.name} не найден столбец 'ID ролика'.")
    id_col, header_row = found

    blocks = []
    cur_ids = []
    cur_chron = {}      # {id: хрон из столбца E} для текущего блока
    cur_time = None
    pending_itogo = None

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        a_val = row[0] if len(row) > 0 else None
        d_val = row[3] if len(row) > 3 else None
        e_val = row[4] if len(row) > 4 else None
        id_val = row[id_col - 1] if len(row) >= id_col else None

        # нормализуем ID
        fid = None
        if isinstance(id_val, int):
            fid = id_val
        elif isinstance(id_val, float) and id_val.is_integer():
            fid = int(id_val)
        elif isinstance(id_val, str) and id_val.strip().isdigit():
            fid = int(id_val.strip())

        if fid is not None:
            cur_ids.append(fid)
            # хрон ролика из столбца E (key-value по id; один хрон на id)
            if isinstance(e_val, (int, float)):
                cur_chron[fid] = int(e_val)
            if cur_time is None:
                if isinstance(a_val, _dt.time):
                    cur_time = a_val.strftime("%H:%M")
                elif a_val not in (None, ""):
                    cur_time = str(a_val).strip()
        else:
            # строка без ID — возможно ИТОГО или пустая
            if isinstance(d_val, str) and d_val.strip().upper() == "ИТОГО":
                if isinstance(e_val, (int, float)):
                    pending_itogo = int(e_val)
            # завершение блока: есть накопленные ролики и встретили разрыв
            if cur_ids and (id_val is None or id_val == ""):
                # закрываем блок только когда блок реально закончился:
                # ориентир — пустая строка ПОСЛЕ итого. Но проще: копим и закрываем
                # на первой пустой строке, где нет ни ID, ни ИТОГО.
                is_blank = (a_val in (None, "")) and (d_val in (None, "")) and (id_val in (None, ""))
                if is_blank:
                    blocks.append({"time": cur_time, "ids": cur_ids,
                                   "chron": cur_chron, "itogo": pending_itogo})
                    cur_ids = []
                    cur_chron = {}
                    cur_time = None
                    pending_itogo = None
    if cur_ids:
        blocks.append({"time": cur_time, "ids": cur_ids,
                       "chron": cur_chron, "itogo": pending_itogo})
    wb.close()
    return blocks



def time_to_filepart(t: str) -> str:
    """'05:25' -> '05-25'. Если время None — 'unknown'."""
    if not t:
        return "unknown"
    return t.replace(":", "-")



def ddmmyy_to_dashed(ddmmyy: str) -> str:
    """'170626' -> '17-06-26'."""
    return f"{ddmmyy[0:2]}-{ddmmyy[2:4]}-{ddmmyy[4:6]}"


# ---------- проверка наличия файлов (обработчик 1) ----------


def check_all_files_exist(blocks: list, src_dir: Path) -> list:
    """Возвращает список (block_index, missing_id). Пусто = всё на месте."""
    missing = []
    for bi, b in enumerate(blocks):
        for fid in b["ids"]:
            if not (src_dir / f"{fid}{EXT}").is_file():
                missing.append((bi, fid))
    return missing


# ---------- построение ffmpeg-команд ----------
